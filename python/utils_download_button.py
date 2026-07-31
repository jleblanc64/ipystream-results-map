import base64
import tempfile
import pandas as pd
from ipywidgets import HTML
from pandas import DataFrame
from openpyxl.utils import get_column_letter

html_template = """<html>
<head>
<meta name="viewport" content="width=device-width, initial-scale=1">
</head>
<body>
<a download="{filename}" href="data:text/csv;base64,{payload}" download>
<button class="p-Widget jupyter-widgets jupyter-button widget-button mod-warning">Download File</button>
</a>
</body>
</html>
"""

def download_button(df: DataFrame) -> HTML:
    sheet_name = "0"
    bytes = None
    with tempfile.NamedTemporaryFile(suffix=".xlsx") as f:
        path = f.name

        # adjust columns size
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            worksheet = writer.sheets[sheet_name]

            for column_cells in worksheet.iter_cols():
                if column_cells:
                    col_idx = column_cells[0].column
                    new_column_letter = get_column_letter(col_idx)

                    header_text = str(df.columns[col_idx - 1]) if col_idx - 1 < len(df.columns) else ""

                    max_length = len(header_text)
                    for cell in column_cells:
                        if cell.value is not None:
                            max_length = max(max_length, len(str(cell.value)))

                    adjusted_width = max_length * 1.23
                    if adjusted_width > 0:
                        worksheet.column_dimensions[new_column_letter].width = adjusted_width

        with open(path, "rb") as file:
            bytes = file.read()

    payload = base64.b64encode(bytes).decode()
    html_button = html_template.format(payload=payload, filename="results.xlsx")
    return HTML(html_button)